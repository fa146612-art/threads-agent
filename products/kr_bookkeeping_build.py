#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
1인 사업자·프리랜서 장부 (한국판)

번역본이 아니라 한국 세법에 맞춰 새로 만든 것.
- 프리랜서 3.3% 원천징수 반영
- 간이/일반 과세 구분과 부가세 예상
- 종합소득세 구간(누진공제) 자동 계산
- 신고 일정 캘린더

세율·기준 검증일 2026-07-28 (본문에 출처 표기)
"""
import sys, datetime
sys.path.insert(0, "/home/claude/products")
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

F = "맑은 고딕"
A, AD, ALT = "2E4A6B", "1E3350", "E3EAF2"
INK, MUTED, LINE = "1F2933", "6B7684", "D9D2C7"
INPUT_BG, WHITE, ACC = "FFF9E6", "FFFFFF", "C1743F"
WON = '#,##0"원";[Red]-#,##0"원";"-"'
PCT = '0.0%'
DATEF = 'yyyy-mm-dd'
thin = Side("thin", color=LINE)
BOX = Border(thin, thin, thin, thin)

def fnt(sz=10, b=False, c=INK, i=False): return Font(name=F, size=sz, bold=b, color=c, italic=i)
def fil(c): return PatternFill("solid", fgColor=c)

def title(ws, t, sub):
    ws["B2"] = t; ws["B2"].font = fnt(18, True, A)
    ws["B3"] = sub; ws["B3"].font = fnt(9, c=MUTED, i=True)
    ws.row_dimensions[2].height = 28; ws.row_dimensions[4].height = 8

def section(ws, row, col, text, span=4):
    for k in range(col, col+span):
        c = ws.cell(row=row, column=k); c.fill = fil(A); c.font = fnt(11, True, WHITE)
    ws.cell(row=row, column=col, value=text); ws.row_dimensions[row].height = 20

def colhdr(ws, row, col, labels):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=col+i, value=lab)
        c.font = fnt(9, True, A); c.fill = fil(ALT); c.border = BOX
        c.alignment = Alignment("center", "center", wrap_text=True)
    ws.row_dimensions[row].height = 26

def inp(c, fmt=None):
    c.fill = fil(INPUT_BG); c.font = fnt(10, c="0000FF"); c.border = BOX
    if fmt: c.number_format = fmt
    return c

def calc(c, fmt=None, b=False, color=INK):
    c.font = fnt(10, b, color); c.border = BOX
    if fmt: c.number_format = fmt
    return c

def widths(ws, spec):
    for k, w in spec.items(): ws.column_dimensions[k].width = w

wb = openpyxl.Workbook()
ROWS = 500          # 입력 여유 행

# ══════════════════════════════════════════════════ 1. 시작하기
ws = wb.active; ws.title = "시작하기"
widths(ws, {"A": 3, "B": 26, "C": 88})
title(ws, "1인 사업자 · 프리랜서 장부",
      "매출과 비용만 넣으면 손익 · 부가세 · 종합소득세가 자동으로 나옵니다.")

blocks = [
    ("이 파일이 하는 일", None),
    (None, "입력하는 곳은 [매출]과 [비용] 두 군데뿐입니다. 나머지는 전부 계산됩니다."),
    (None, "회계 지식이 없어도 됩니다. 월 구독료도 없습니다. 데이터는 본인 컴퓨터에만 남습니다."),
    ("시트 구성", None),
    ("1. 기본설정", "상호, 과세 유형(일반/간이), 업종, 카테고리 목록. 처음 한 번 5분."),
    ("2. 매출", "받은 돈을 기록합니다. 프리랜서는 3.3% 원천징수 여부를 체크하면 자동 반영됩니다."),
    ("3. 비용", "쓴 돈을 기록합니다. 경비 인정 여부와 매입세액 공제 여부를 표시합니다."),
    ("4. 대시보드", "이번 달 매출·비용·순이익·이익률. 월을 바꿔가며 볼 수 있습니다."),
    ("5. 손익계산", "12개월 전체와 연간 합계. 그대로 손익계산서로 씁니다."),
    ("6. 세금요약", "부가세 예상액, 종합소득세 예상액, 원천징수 정산 예상, 신고 일정."),
    ("7. 거래처", "누가 얼마 줬는지, 아직 못 받은 돈이 얼마인지."),
    ("색깔 규칙", None),
    ("노란 칸 · 파란 글씨", "여기에 입력하세요."),
    ("흰 칸 · 검은 글씨", "자동 계산됩니다. 덮어쓰면 수식이 지워집니다."),
    ("시작 순서", None),
    ("1단계", "[기본설정]에서 상호와 과세 유형을 고릅니다. 모르면 홈택스 > 사업자등록 내역에서 확인."),
    ("2단계", "카테고리 목록을 본인 업종에 맞게 고칩니다."),
    ("3단계", "[매출]과 [비용]의 예시 행을 지우고 실제 내역을 넣기 시작합니다."),
    ("꼭 읽어주세요", None),
    (None, "이 파일은 기록을 정리하고 세금을 '예상'해 주는 도구입니다. 세무 대리나 신고 대행이 아닙니다."),
    (None, "실제 신고 금액은 공제·감면·업종에 따라 달라집니다. 신고 전 세무사 또는 홈택스에서 확인하세요."),
    (None, "세율 기준일: 2026-07-28 (국세청 종합소득세 세율표 / 간이과세 기준 1억 400만원)"),
]
r = 6
for label, text in blocks:
    if text is None:
        section(ws, r, 2, label, span=2)
    else:
        if label: ws.cell(row=r, column=2, value=label).font = fnt(10, True)
        ws.cell(row=r, column=3, value=text).font = fnt(10, c=INK if label else MUTED)
    r += 1

# ══════════════════════════════════════════════════ 2. 기본설정
st = wb.create_sheet("기본설정")
widths(st, {"A": 3, "B": 22, "C": 24, "D": 34, "E": 3, "F": 24, "G": 24, "H": 20})
title(st, "기본설정", "처음 한 번만 채우면 됩니다.")

section(st, 6, 2, "사업 정보", span=3)
rows = [("상호 / 이름", "내 사업"), ("사업 형태", "프리랜서"), ("과세 유형", "간이과세자"),
        ("업종", "서비스"), ("사업 시작 연도", 2026)]
for i, (k, v) in enumerate(rows):
    st.cell(row=7+i, column=2, value=k).font = fnt(10, True)
    inp(st.cell(row=7+i, column=3, value=v))
st["D7"] = "세금계산·신고안내에 쓰입니다"; st["D7"].font = fnt(9, c=MUTED)
st["D8"] = "프리랜서 / 개인사업자"; st["D8"].font = fnt(9, c=MUTED)
st["D9"] = "홈택스 > 사업자등록 내역에서 확인"; st["D9"].font = fnt(9, c=MUTED)
st["D10"] = "간이과세 부가율 판단에 쓰입니다"; st["D10"].font = fnt(9, c=MUTED)

dv_form = DataValidation(type="list", formula1='"프리랜서,개인사업자"', allow_blank=True)
dv_tax = DataValidation(type="list", formula1='"간이과세자,일반과세자,면세사업자"', allow_blank=True)
dv_ind = DataValidation(type="list", formula1='"서비스,소매,음식점,숙박,제조,건설,운수,기타"', allow_blank=True)
for dv in (dv_form, dv_tax, dv_ind): st.add_data_validation(dv)
dv_form.add(st["C8"]); dv_tax.add(st["C9"]); dv_ind.add(st["C10"])

section(st, 13, 2, "세금 설정", span=3)
tx = [("원천징수율 (프리랜서)", 0.033, "지급액에서 떼는 비율. 소득세 3% + 지방세 0.3%"),
      ("부가세율 (일반과세)", 0.10, "일반과세자 매출세액 계산용"),
      ("간이 부가가치율", 0.15, "업종별 15~40%. 서비스업 통상 15~30%"),
      ("차량 km당 단가", 300, "차량 사용 기록용. 본인 기준으로 조정")]
for i, (k, v, note) in enumerate(tx):
    st.cell(row=14+i, column=2, value=k).font = fnt(10, True)
    inp(st.cell(row=14+i, column=3, value=v), PCT if v < 1 else WON)
    st.cell(row=14+i, column=4, value=note).font = fnt(9, c=MUTED)

section(st, 6, 6, "매출 카테고리", span=1)
inc_cats = ["용역/서비스", "제품 판매", "온라인 판매", "강의/컨설팅", "광고/제휴", "기타 수입"]
for i, c in enumerate(inc_cats): inp(st.cell(row=7+i, column=6, value=c))

section(st, 14, 6, "비용 카테고리", span=1)
exp_cats = ["재료비/매입", "임차료", "인건비", "통신비", "광고선전비", "차량유지비",
            "소모품비", "지급수수료", "접대비", "여비교통비", "보험료", "세금과공과", "기타"]
for i, c in enumerate(exp_cats): inp(st.cell(row=15+i, column=6, value=c))

section(st, 6, 7, "결제 수단", span=1)
for i, c in enumerate(["계좌이체", "카드", "현금", "간편결제", "기타"]):
    inp(st.cell(row=7+i, column=7, value=c))

st["H6"] = "간이과세 참고"; st["H6"].font = fnt(10, True, A)
notes = ["연매출 1억 400만원 미만 → 간이과세",
         "연매출 4,800만원 미만 → 부가세 납부면제",
         "(신고는 해야 함)",
         "간이 신고: 매년 1월 25일",
         "일반 신고: 1월 · 7월",
         "종합소득세: 매년 5월",
         "",
         "제조·도매·전문직 등은",
         "간이과세 배제 업종입니다."]
for i, n in enumerate(notes):
    st.cell(row=7+i, column=8, value=n).font = fnt(9, c=MUTED)

# ══════════════════════════════════════════════════ 3. 매출
inc = wb.create_sheet("매출")
widths(inc, {"A": 3, "B": 12, "C": 22, "D": 18, "E": 14, "F": 13, "G": 11, "H": 13, "I": 13, "J": 10, "K": 15, "L": 22})
title(inc, "매출", "받은 돈을 기록합니다. 노란 칸만 채우면 됩니다.")
colhdr(inc, 6, 2, ["날짜", "거래처", "내용", "카테고리",
                   "공급가액", "부가세", "총 입금액", "원천징수 여부", "입금여부",
                   "원천징수액(3.3%)", "메모"])
dv_inc = DataValidation(type="list", formula1="=기본설정!$F$7:$F$12", allow_blank=True)
dv_yn = DataValidation(type="list", formula1='"예,아니오"', allow_blank=True)
inc.add_data_validation(dv_inc); inc.add_data_validation(dv_yn)

samples = [(datetime.date(2026,1,15), "A거래처", "1월 용역", "용역/서비스", 1000000, "아니오", "예"),
           (datetime.date(2026,1,28), "B스토어", "제품 판매", "제품 판매", 350000, "예", "예")]
for i in range(ROWS):
    r = 7+i
    if i < len(samples):
        d, cl, memo, cat, amt, wh, paid = samples[i]
        inp(inc.cell(row=r, column=2, value=d), DATEF)
        inp(inc.cell(row=r, column=3, value=cl)); inp(inc.cell(row=r, column=4, value=memo))
        inp(inc.cell(row=r, column=5, value=cat)); inp(inc.cell(row=r, column=6, value=amt), WON)
        inp(inc.cell(row=r, column=9, value=wh)); inp(inc.cell(row=r, column=10, value=paid))
    else:
        inp(inc.cell(row=r, column=2), DATEF)
        for col in (3, 4, 5, 9, 10): inp(inc.cell(row=r, column=col))
        inp(inc.cell(row=r, column=6), WON)
    # 부가세 = 일반과세자만 매출세액 발생
    calc(inc.cell(row=r, column=7,
         value=f'=IF($F{r}="","",IF(기본설정!$C$9="일반과세자",ROUND($F{r}*기본설정!$C$15,0),0))'), WON)
    calc(inc.cell(row=r, column=8, value=f'=IF($F{r}="","",$F{r}+$G{r})'), WON, b=True)
    calc(inc.cell(row=r, column=11,
         value=f'=IF($F{r}="","",IF($I{r}="예",ROUND($F{r}*기본설정!$C$14,0),0))'), WON)
    inp(inc.cell(row=r, column=12))
    dv_inc.add(inc.cell(row=r, column=5)); dv_yn.add(inc.cell(row=r, column=9)); dv_yn.add(inc.cell(row=r, column=10))
inc.freeze_panes = "B7"

# ══════════════════════════════════════════════════ 4. 비용
exp = wb.create_sheet("비용")
widths(exp, {"A": 3, "B": 12, "C": 22, "D": 18, "E": 14, "F": 13, "G": 11, "H": 13, "I": 12, "J": 12, "K": 20})
title(exp, "비용", "쓴 돈을 기록합니다. 경비 인정 여부를 꼭 표시하세요.")
colhdr(exp, 6, 2, ["날짜", "거래처", "내용", "카테고리",
                   "금액(부가세포함)", "공급가액", "매입세액", "경비인정", "결제수단", "메모"])
dv_exp = DataValidation(type="list", formula1="=기본설정!$F$15:$F$27", allow_blank=True)
dv_pay = DataValidation(type="list", formula1="=기본설정!$G$7:$G$11", allow_blank=True)
dv_yn2 = DataValidation(type="list", formula1='"예,아니오"', allow_blank=True)
for dv in (dv_exp, dv_pay, dv_yn2): exp.add_data_validation(dv)

esamples = [(datetime.date(2026,1,5), "건물주", "1월 임차료", "임차료", 800000, "예", "계좌이체"),
            (datetime.date(2026,1,12), "통신사", "인터넷", "통신비", 44000, "예", "카드")]
for i in range(ROWS):
    r = 7+i
    if i < len(esamples):
        d, v, memo, cat, amt, ded, pay = esamples[i]
        inp(exp.cell(row=r, column=2, value=d), DATEF)
        inp(exp.cell(row=r, column=3, value=v)); inp(exp.cell(row=r, column=4, value=memo))
        inp(exp.cell(row=r, column=5, value=cat)); inp(exp.cell(row=r, column=6, value=amt), WON)
        inp(exp.cell(row=r, column=9, value=ded)); inp(exp.cell(row=r, column=10, value=pay))
    else:
        inp(exp.cell(row=r, column=2), DATEF)
        for col in (3, 4, 5, 9, 10): inp(exp.cell(row=r, column=col))
        inp(exp.cell(row=r, column=6), WON)
    calc(exp.cell(row=r, column=7, value=f'=IF($F{r}="","",ROUND($F{r}/1.1,0))'), WON)
    calc(exp.cell(row=r, column=8,
         value=f'=IF($F{r}="","",IF(기본설정!$C$9="일반과세자",$F{r}-$G{r},0))'), WON)
    dv_exp.add(exp.cell(row=r, column=5)); dv_yn2.add(exp.cell(row=r, column=9)); dv_pay.add(exp.cell(row=r, column=10))
exp.freeze_panes = "B7"

IR, ER = f"7:{6+ROWS}", f"7:{6+ROWS}"
INC_D, INC_S, INC_W = f"매출!$B$7:$B${6+ROWS}", f"매출!$F$7:$F${6+ROWS}", f"매출!$K$7:$K${6+ROWS}"
INC_C = f"매출!$E$7:$E${6+ROWS}"
EXP_D, EXP_A, EXP_V = f"비용!$B$7:$B${6+ROWS}", f"비용!$F$7:$F${6+ROWS}", f"비용!$H$7:$H${6+ROWS}"
EXP_C, EXP_DED = f"비용!$E$7:$E${6+ROWS}", f"비용!$I$7:$I${6+ROWS}"

# ══════════════════════════════════════════════════ 5. 손익계산
pl = wb.create_sheet("손익계산")
widths(pl, {"A": 3, "B": 22})
for i in range(13): pl.column_dimensions[get_column_letter(3+i)].width = 15
title(pl, "손익계산", "12개월 전체와 연간 합계입니다.")
pl["B6"] = "연도"; pl["B6"].font = fnt(10, True); inp(pl["C6"], None).value = 2026

hdr = ["항목"] + [f"{m}월" for m in range(1, 13)] + ["연간"]
colhdr(pl, 8, 2, hdr)

def m_first(m): return f'DATE($C$6,{m},1)'
def m_last(m): return f'EOMONTH(DATE($C$6,{m},1),0)'

lines = [
    ("매출액", lambda m: f'=SUMIFS({INC_S},{INC_D},">="&{m_first(m)},{INC_D},"<="&{m_last(m)})'),
    ("비용 합계", lambda m: f'=SUMIFS({EXP_A},{EXP_D},">="&{m_first(m)},{EXP_D},"<="&{m_last(m)})'),
    ("  경비 인정분", lambda m: f'=SUMIFS({EXP_A},{EXP_D},">="&{m_first(m)},{EXP_D},"<="&{m_last(m)},{EXP_DED},"예")'),
    ("순이익", None),
    ("이익률", None),
    ("원천징수액", lambda m: f'=SUMIFS({INC_W},{INC_D},">="&{m_first(m)},{INC_D},"<="&{m_last(m)})'),
]
row = 9
row_map = {}
for name, fx in lines:
    pl.cell(row=row, column=2, value=name).font = fnt(10, True if not name.startswith("  ") else False)
    row_map[name.strip()] = row
    for m in range(1, 13):
        col = 2+m
        if fx:
            calc(pl.cell(row=row, column=col, value=fx(m)), WON)
        row += 0
    row += 1

R_SALES, R_EXP, R_DED, R_NET, R_MGN, R_WH = (row_map["매출액"], row_map["비용 합계"],
    row_map["경비 인정분"], row_map["순이익"], row_map["이익률"], row_map["원천징수액"])
for m in range(1, 13):
    col = get_column_letter(2+m)
    calc(pl[f"{col}{R_NET}"], WON, b=True).value = f"={col}{R_SALES}-{col}{R_EXP}"
    calc(pl[f"{col}{R_MGN}"], PCT).value = f'=IF({col}{R_SALES}=0,"",{col}{R_NET}/{col}{R_SALES})'
for rr in (R_SALES, R_EXP, R_DED, R_NET, R_WH):
    calc(pl.cell(row=rr, column=15, value=f"=SUM(C{rr}:N{rr})"), WON, b=True)
calc(pl.cell(row=R_MGN, column=15, value=f'=IF(O{R_SALES}=0,"",O{R_NET}/O{R_SALES})'), PCT, b=True)

section(pl, R_WH+2, 2, "카테고리별 비용 (연간)", span=3)
rr = R_WH+3
colhdr(pl, rr, 2, ["카테고리", "금액", "비중"])
for i in range(13):
    r2 = rr+1+i
    calc(pl.cell(row=r2, column=2, value=f"=기본설정!F{15+i}"))
    calc(pl.cell(row=r2, column=3,
        value=f'=IF($B{r2}="","",SUMIFS({EXP_A},{EXP_C},$B{r2},{EXP_D},">="&DATE($C$6,1,1),{EXP_D},"<="&DATE($C$6,12,31)))'), WON)
    calc(pl.cell(row=r2, column=4, value=f'=IF(OR($B{r2}="",$O${R_EXP}=0),"",C{r2}/$O${R_EXP})'), PCT)

# ══════════════════════════════════════════════════ 6. 대시보드
db = wb.create_sheet("대시보드")
widths(db, {"A": 3, "B": 24, "C": 20, "D": 6, "E": 24, "F": 20, "G": 6, "H": 26, "I": 20})
title(db, "대시보드", "월을 바꾸면 그 달 숫자가 나옵니다.")
db["B6"] = "보고 싶은 월"; db["B6"].font = fnt(10, True)
inp(db["C6"]).value = 1
dv_m = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10,11,12"', allow_blank=True)
db.add_data_validation(dv_m); dv_m.add(db["C6"])
db["D6"] = "월"; db["D6"].font = fnt(10)

def card(ws, r, c, label, formula, fmt=WON, color=A):
    ws.cell(row=r, column=c, value=label).font = fnt(10, True, MUTED)
    cell = ws.cell(row=r+1, column=c, value=formula)
    cell.font = fnt(20, True, color); cell.number_format = fmt

section(db, 8, 2, "이번 달", span=8)
IDX = f"INDEX(손익계산!$C${R_SALES}:$N${R_SALES},$C$6)"
card(db, 10, 2, "매출", f"=INDEX(손익계산!$C${R_SALES}:$N${R_SALES},$C$6)")
card(db, 10, 5, "비용", f"=INDEX(손익계산!$C${R_EXP}:$N${R_EXP},$C$6)", color=ACC)
card(db, 10, 8, "순이익", f"=INDEX(손익계산!$C${R_NET}:$N${R_NET},$C$6)")
card(db, 13, 2, "이익률", f"=INDEX(손익계산!$C${R_MGN}:$N${R_MGN},$C$6)", PCT)
card(db, 13, 5, "원천징수액", f"=INDEX(손익계산!$C${R_WH}:$N${R_WH},$C$6)", color=ACC)
card(db, 13, 8, "경비 인정분", f"=INDEX(손익계산!$C${R_DED}:$N${R_DED},$C$6)")

section(db, 17, 2, "연간 누계", span=8)
card(db, 19, 2, "연 매출", f"=손익계산!$O${R_SALES}")
card(db, 19, 5, "연 비용", f"=손익계산!$O${R_EXP}", color=ACC)
card(db, 19, 8, "연 순이익", f"=손익계산!$O${R_NET}")

db["B23"] = "과세 유형 확인"; db["B23"].font = fnt(10, True, A)
db["B24"] = (f'=IF(손익계산!$O${R_SALES}>=104000000,'
             f'"연매출 1억 400만원 이상 → 다음 해 7월 1일부터 일반과세자로 전환됩니다.",'
             f'IF(손익계산!$O${R_SALES}<48000000,'
             f'"연매출 4,800만원 미만 → 간이과세자라면 부가세 납부는 면제됩니다. 신고는 해야 합니다.",'
             f'"간이과세 구간입니다. (1억 400만원 미만)"))')
db["B24"].font = fnt(10, c=INK)

# ══════════════════════════════════════════════════ 7. 세금요약
tx = wb.create_sheet("세금요약")
widths(tx, {"A": 3, "B": 30, "C": 20, "D": 46})
title(tx, "세금 요약", "예상액입니다. 실제 신고액은 공제·감면에 따라 달라집니다.")

section(tx, 6, 2, "부가가치세 (예상)", span=3)
tx["B7"] = "매출세액"; calc(tx["C7"], WON).value = f'=IF(기본설정!$C$9="일반과세자",SUM(매출!$G$7:$G${6+ROWS}),0)'
tx["D7"] = "일반과세자만 발생"; tx["D7"].font = fnt(9, c=MUTED)
tx["B8"] = "매입세액"; calc(tx["C8"], WON).value = f'=IF(기본설정!$C$9="일반과세자",SUM(비용!$H$7:$H${6+ROWS}),0)'
tx["D8"] = "적격증빙(세금계산서·카드) 있는 것만 실제 공제"; tx["D8"].font = fnt(9, c=MUTED)
tx["B9"] = "일반과세 납부예상"; calc(tx["C9"], WON, b=True).value = "=MAX(C7-C8,0)"
tx["B10"] = "간이과세 납부예상"
calc(tx["C10"], WON, b=True).value = (f'=IF(기본설정!$C$9="간이과세자",'
    f'IF(손익계산!$O${R_SALES}<48000000,0,ROUND(손익계산!$O${R_SALES}*기본설정!$C$16*0.1,0)),0)')
tx["D10"] = "매출 × 업종별 부가가치율 × 10% / 4,800만원 미만은 납부면제"; tx["D10"].font = fnt(9, c=MUTED)

section(tx, 12, 2, "종합소득세 (예상)", span=3)
tx["B13"] = "사업소득 (순이익)"; calc(tx["C13"], WON).value = f"=손익계산!$O${R_NET}"
tx["B14"] = "과세표준 (가정)"; calc(tx["C14"], WON).value = "=MAX(C13-1500000,0)"
tx["D14"] = "기본공제 150만원만 반영한 단순 가정. 실제는 인적·연금 등 추가 공제 있음"; tx["D14"].font = fnt(9, c=MUTED)
tx["B15"] = "산출세액"
calc(tx["C15"], WON, b=True).value = (
 '=ROUND(IF(C14<=14000000,C14*0.06,'
 'IF(C14<=50000000,C14*0.15-1260000,'
 'IF(C14<=88000000,C14*0.24-5760000,'
 'IF(C14<=150000000,C14*0.35-15440000,'
 'IF(C14<=300000000,C14*0.38-19940000,'
 'IF(C14<=500000000,C14*0.40-25940000,'
 'IF(C14<=1000000000,C14*0.42-35940000,C14*0.45-65940000))))))),0)')
tx["D15"] = "국세청 세율표 기준 (6%~45%, 누진공제 반영)"; tx["D15"].font = fnt(9, c=MUTED)
tx["B16"] = "지방소득세 (10%)"; calc(tx["C16"], WON).value = "=ROUND(C15*0.1,0)"
tx["B17"] = "이미 낸 원천징수액"; calc(tx["C17"], WON).value = f"=손익계산!$O${R_WH}"
tx["B18"] = "정산 예상"; calc(tx["C18"], WON, b=True).value = "=C15+C16-C17"
tx["D18"] = "음수면 환급 예상, 양수면 추가 납부 예상"; tx["D18"].font = fnt(9, c=MUTED)
tx.conditional_formatting.add("C18",
    CellIsRule(operator="lessThan", formula=["0"], font=Font(name=F, size=10, bold=True, color="0E9F6E")))
tx.conditional_formatting.add("C18",
    CellIsRule(operator="greaterThan", formula=["0"], font=Font(name=F, size=10, bold=True, color="D64545")))

section(tx, 20, 2, "신고 일정", span=3)
sched = [("1월 25일", "부가세 확정신고", "간이과세자 연 1회 / 일반과세자 2기 확정"),
         ("5월 1일 ~ 6월 1일", "종합소득세 신고", "전년도 소득 대상. 프리랜서 환급도 이때"),
         ("7월 25일", "부가세 확정신고", "일반과세자 1기 확정"),
         ("수시", "지원금 공고 확인", "마감 시각이 16시인 공고가 많습니다")]
colhdr(tx, 21, 2, ["시기", "할 일", "메모"])
for i, (a1, b1, c1) in enumerate(sched):
    r2 = 22+i
    calc(tx.cell(row=r2, column=2, value=a1), b=True)
    calc(tx.cell(row=r2, column=3, value=b1))
    calc(tx.cell(row=r2, column=4, value=c1))

tx["B28"] = "출처 및 기준일"; tx["B28"].font = fnt(10, True, A)
for i, s in enumerate([
    "종합소득세 세율: 국세청 종합소득세 세율표 (1,400만원 이하 6% ~ 10억 초과 45%)",
    "간이과세 기준: 연매출 1억 400만원 미만 / 납부면제 4,800만원 미만",
    "종합소득세 신고: 2026년 5월 1일 ~ 6월 1일",
    "확인일: 2026-07-28. 세법은 바뀝니다. 신고 전 홈택스에서 다시 확인하세요."]):
    tx.cell(row=29+i, column=2, value=s).font = fnt(9, c=MUTED)

# ══════════════════════════════════════════════════ 8. 거래처
cl = wb.create_sheet("거래처")
widths(cl, {"A": 3, "B": 26, "C": 18, "D": 18, "E": 18, "F": 14})
title(cl, "거래처", "누가 얼마 줬는지, 아직 못 받은 돈이 얼마인지.")
colhdr(cl, 6, 2, ["거래처명", "총 매출", "입금 완료", "미수금", "건수"])
for i in range(40):
    r2 = 7+i
    inp(cl.cell(row=r2, column=2))
    calc(cl.cell(row=r2, column=3,
        value=f'=IF($B{r2}="","",SUMIF(매출!$C$7:$C${6+ROWS},$B{r2},매출!$F$7:$F${6+ROWS}))'), WON)
    calc(cl.cell(row=r2, column=4,
        value=f'=IF($B{r2}="","",SUMIFS(매출!$F$7:$F${6+ROWS},매출!$C$7:$C${6+ROWS},$B{r2},매출!$J$7:$J${6+ROWS},"예"))'), WON)
    calc(cl.cell(row=r2, column=5, value=f'=IF($B{r2}="","",C{r2}-D{r2})'), WON, b=True)
    calc(cl.cell(row=r2, column=6,
        value=f'=IF($B{r2}="","",COUNTIF(매출!$C$7:$C${6+ROWS},$B{r2}))'))
cl.conditional_formatting.add(f"E7:E46",
    CellIsRule(operator="greaterThan", formula=["0"], font=Font(name=F, size=10, bold=True, color="D64545")))

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

out = "/home/claude/products/04_kr_bookkeeping/1인사업자_프리랜서_장부.xlsx"
wb.save(out)
print("저장:", out)

# 수식 개수 세기
n = 0
wb2 = openpyxl.load_workbook(out)
for s in wb2.worksheets:
    for row in s.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="): n += 1
print("시트", len(wb2.worksheets), "개 / 수식", n, "개")
